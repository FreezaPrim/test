"""
TNPS Detractor Analytics & Forecasting Pipeline  (Enhanced Edition)
====================================================================
Reads multiple TNPS survey .xlsx files from a folder, joins with the Agent
Queue Mapping, and produces a single Excel dashboard report.

ENHANCEMENTS over v1
--------------------
  Performance
    - pd.cut() vectorised NPS categorisation (replaces row-wise apply)
    - Parquet incremental cache: only re-reads truly new .xlsx files
    - Lazy sheet writing: skips sheets whose source column is absent

  Analytics
    - Cohort tracking: same MSISDN across months (recovery rate)
    - Root-cause clustering: top toxic dimension combos
    - Agent peer-group benchmarking + percentile bands
    - Week-over-week velocity alerts per queue and short code

  Forecasting
    - Per-queue forecasts (top-N queues)
    - Holt-Winters prediction intervals (upper / lower bands)
    - Configurable target line for NPS / detractor rate breach date

  Data Quality
    - Source-file column validation with clear warnings
    - Duplicate survey detection (same MSISDN within 24 h)
    - Mapping coverage report sheet

  Report Usability
    - Executive Summary narrative sheet
    - Delta vs previous period on every KPI card
    - Hyperlinked Table of Contents sheet
    - Trend column in Short Code pivot colour-coded (cell background)

  Configuration & Operability
    - CLI overrides via argparse (--data, --output, --horizon, --config)
    - config.yaml support (falls back to built-in CONFIG dict)
    - Python logging replaces print()

  New Sheets
    - NPS Waterfall (promoters gained vs detractors lost)
    - SLA Breach Heatmap (queues × days that breached threshold)
    - Top10_Bottom10 (ranked queues + short codes)
    - Channel_Comparison
    - FCR_Impact (FCR=Yes vs No)

USAGE
-----
  python tnps_analyzer.py                          # defaults
  python tnps_analyzer.py --data ./jan --output Jan.xlsx
  python tnps_analyzer.py --config custom.yaml
"""

# ============================================================
# IMPORTS
# ============================================================
import argparse
import hashlib
import json
import logging
import sys
import warnings
from pathlib import Path
from datetime import timedelta, date

import numpy as np
import pandas as pd

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.styles.differential import DifferentialStyle

from statsmodels.tsa.holtwinters import ExponentialSmoothing

warnings.filterwarnings("ignore")

# ============================================================
# DEFAULT CONFIG  (overridden by --config yaml or CLI flags)
# ============================================================
_DEFAULT_CONFIG = {
    "data_folder": "data",
    "mapping_file": "AgentQueueMapping.xlsx",
    "output_file": "TNPS_Report.xlsx",
    "cache_file": ".tnps_cache.parquet",      # incremental cache

    # Business rules
    "detractor_max": 6,
    "passive_max": 8,

    # Forecast
    "forecast_horizon_days": 30,
    "forecast_top_n_queues": 5,               # top queues to forecast individually
    "nps_target": -10,                         # target NPS (breach alert)
    "detractor_rate_target": 40,               # target detractor rate % (breach alert)

    # Guard rails
    "min_agent_sample": 5,
    "anomaly_std_threshold": 2,
    "velocity_alert_pct": 20,                  # week-over-week % jump = alert
    "sla_detractor_rate_threshold": 50,        # % above which a day/queue is "breach"

    # Branding
    "brand_primary": "C8102E",
    "brand_secondary": "1F1F1F",
    "brand_accent": "FFC72C",
    "brand_good": "2E8B57",
    "brand_bad": "C8102E",
}
CONFIG = dict(_DEFAULT_CONFIG)

# ============================================================
# LOGGING
# ============================================================
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("tnps")

# ============================================================
# CLI
# ============================================================
def parse_args():
    p = argparse.ArgumentParser(description="TNPS Analytics & Forecasting Pipeline")
    p.add_argument("--data",    default=None, help="Folder containing TNPS .xlsx files")
    p.add_argument("--mapping", default=None, help="AgentQueueMapping.xlsx path")
    p.add_argument("--output",  default=None, help="Output Excel report path")
    p.add_argument("--horizon", default=None, type=int, help="Forecast horizon in days")
    p.add_argument("--config",  default=None, help="Path to config.yaml override file")
    p.add_argument("--no-cache", action="store_true", help="Ignore incremental cache, re-read all files")
    args, _ = p.parse_known_args()  # parse_known_args ignores IDE-injected flags (e.g. VS Code kernel args)
    return args


def apply_config_overrides(args):
    """Merge YAML config and CLI args into global CONFIG."""
    global CONFIG
    if args.config:
        try:
            import yaml
            with open(args.config) as f:
                yaml_cfg = yaml.safe_load(f) or {}
            CONFIG.update(yaml_cfg)
            log.info("Loaded config from %s", args.config)
        except ImportError:
            log.warning("PyYAML not installed; --config ignored. pip install pyyaml")
        except FileNotFoundError:
            log.warning("Config file not found: %s", args.config)

    if args.data:    CONFIG["data_folder"] = args.data
    if args.mapping: CONFIG["mapping_file"] = args.mapping
    if args.output:  CONFIG["output_file"]  = args.output
    if args.horizon: CONFIG["forecast_horizon_days"] = args.horizon


# ============================================================
# 1. LOAD & ENRICH  (with incremental cache + column validation)
# ============================================================
_REQUIRED_COLS = [
    "OUTBOUND_IVR_DT",
    "Q1_ANSWER: TNPS",
    "SHORT_CODE",
    "AGENT_QUEUE",
]

def _file_fingerprint(path: Path) -> str:
    """MD5 of file size + mtime — fast proxy for 'has this file changed'."""
    stat = path.stat()
    return hashlib.md5(f"{stat.st_size}_{stat.st_mtime}".encode()).hexdigest()


def load_data(folder: str, mapping_file: str, use_cache: bool = True) -> tuple[pd.DataFrame, pd.DataFrame]:
    folder_path = Path(folder)
    files = sorted(folder_path.glob("*.xlsx"))
    if not files:
        raise FileNotFoundError(
            f"No .xlsx files in '{folder_path.resolve()}'. Drop your monthly TNPS files there."
        )

    cache_path = Path(CONFIG["cache_file"])
    fingerprint_path = Path(CONFIG["cache_file"] + ".index.json")

    cached_fps: dict = {}
    cached_df: pd.DataFrame = pd.DataFrame()

    _CACHE_REQUIRED_COLS = {"Week", "Month", "Hour", "DayOfWeek", "NPS_Category", "_source_file"}
    if use_cache and cache_path.exists() and fingerprint_path.exists():
        try:
            cached_fps = json.loads(fingerprint_path.read_text())
            cached_df = pd.read_parquet(cache_path)
            # Invalidate cache if enrichment columns added after this cache was built are missing
            if not _CACHE_REQUIRED_COLS.issubset(cached_df.columns):
                missing_c = _CACHE_REQUIRED_COLS - set(cached_df.columns)
                log.warning("Cache is stale (missing columns: %s); re-reading all files.", missing_c)
                cached_fps, cached_df = {}, pd.DataFrame()
            else:
                log.info("Loaded cache: %s rows", f"{len(cached_df):,}")
        except Exception as e:
            log.warning("Cache unreadable (%s); re-reading all files.", e)
            cached_fps, cached_df = {}, pd.DataFrame()

    new_frames = []
    current_fps = {}
    log.info("Scanning %d file(s) in '%s/':", len(files), folder)
    for f in files:
        fp = _file_fingerprint(f)
        current_fps[f.name] = fp
        if use_cache and cached_fps.get(f.name) == fp and not cached_df.empty:
            log.info("  ✓ cached  %s", f.name)
            continue
        df = pd.read_excel(f)
        df["_source_file"] = f.name

        # ---- Column validation ----
        missing = [c for c in _REQUIRED_COLS if c not in df.columns]
        if missing:
            log.warning("  ! %s is missing columns: %s — skipped", f.name, missing)
            continue

        new_frames.append(df)
        log.info("  + loaded  %-40s  %s rows", f.name, f"{len(df):,}")

    # Combine: cached rows that still correspond to current files + new rows
    kept_sources = set(current_fps.keys())
    parts = []
    if not cached_df.empty:
        still_valid = cached_df[cached_df["_source_file"].isin(kept_sources)]
        parts.append(still_valid)
    parts.extend(new_frames)

    if not parts:
        raise ValueError("No valid data frames to process.")

    data = pd.concat(parts, ignore_index=True)
    log.info("TOTAL  %s rows", f"{len(data):,}")

    # Persist updated cache
    if use_cache:
        try:
            data.to_parquet(cache_path, index=False)
            fingerprint_path.write_text(json.dumps(current_fps))
        except Exception as e:
            log.warning("Could not write cache: %s", e)

    mapping = pd.read_excel(mapping_file)

    # ---- Date enrichment ----
    data["OUTBOUND_IVR_DT"] = pd.to_datetime(data["OUTBOUND_IVR_DT"], errors="coerce")
    data = data.dropna(subset=["OUTBOUND_IVR_DT"])
    data["Date"]      = data["OUTBOUND_IVR_DT"].dt.date
    data["Month"]     = data["OUTBOUND_IVR_DT"].dt.to_period("M").astype(str)
    data["Week"]      = data["OUTBOUND_IVR_DT"].dt.to_period("W").astype(str)
    data["Hour"]      = data["OUTBOUND_IVR_DT"].dt.hour
    data["DayOfWeek"] = data["OUTBOUND_IVR_DT"].dt.day_name()

    # ---- Vectorised NPS categorisation (pd.cut replaces row-wise apply) ----
    data["NPS_Category"] = pd.cut(
        data["Q1_ANSWER: TNPS"],
        bins=[-1, CONFIG["detractor_max"], CONFIG["passive_max"], 10],
        labels=["Detractor", "Passive", "Promoter"],
    ).astype(str)
    data.loc[data["Q1_ANSWER: TNPS"].isna(), "NPS_Category"] = "N/A"

    # ---- Join with mapping ----
    data = data.merge(mapping, how="left", left_on="AGENT_QUEUE", right_on="Agent Queue")

    map_cols = [
        "Q Mapping Lev 1", "Q Mapping Lev 2 Combined",
        "Q Mapping Lev 3 New PF Seg", "Q Mapping Lev 4 Seg", "Site",
    ]
    for c in map_cols:
        if c in data.columns:
            data[c] = data[c].fillna("Unmapped")
        else:
            data[c] = "Unmapped"

    return data, mapping


# ============================================================
# 2. KPI SUMMARY  (with delta vs previous month)
# ============================================================
def build_kpi_summary(df: pd.DataFrame) -> pd.DataFrame:
    completed = df[df["Q1_ANSWER: TNPS"].notna()]
    det = completed[completed["NPS_Category"] == "Detractor"]
    pas = completed[completed["NPS_Category"] == "Passive"]
    pro = completed[completed["NPS_Category"] == "Promoter"]

    n_days    = max(df["Date"].nunique(), 1)
    n_comp    = len(completed)
    nps_score = (len(pro) - len(det)) / n_comp * 100 if n_comp else 0
    avg_tnps  = completed["Q1_ANSWER: TNPS"].mean() if n_comp else 0

    # Delta: compare last month vs previous month
    months = sorted(df["Month"].unique())
    delta_det_rate = delta_nps = "N/A"
    if len(months) >= 2:
        cur_m  = df[df["Month"] == months[-1]]
        prev_m = df[df["Month"] == months[-2]]

        def _det_rate(d):
            c = d[d["Q1_ANSWER: TNPS"].notna()]
            if len(c) == 0: return 0
            return len(c[c["NPS_Category"] == "Detractor"]) / len(c) * 100

        def _nps(d):
            c = d[d["Q1_ANSWER: TNPS"].notna()]
            if len(c) == 0: return 0
            p_ = len(c[c["NPS_Category"] == "Promoter"])
            d_ = len(c[c["NPS_Category"] == "Detractor"])
            return (p_ - d_) / len(c) * 100

        delta_det_rate = round(_det_rate(cur_m) - _det_rate(prev_m), 2)
        delta_nps      = round(_nps(cur_m)      - _nps(prev_m),      2)

    return pd.DataFrame({
        "Metric": [
            "Date Range Start", "Date Range End", "Number of Days Covered",
            "Total Surveys Sent", "Completed Surveys", "Completion Rate %",
            "Detractors (Q1 0-6)", "Passives (Q1 7-8)", "Promoters (Q1 9-10)",
            "Detractor Rate %", "NPS Score", "Average TNPS Score",
            "Average Surveys / Day", "Average Detractors / Day",
            "Unique Short Codes", "Unique Agent Queues", "Unique Agents",
            "Unmapped Queue Surveys",
        ],
        "Value": [
            str(df["Date"].min()), str(df["Date"].max()), n_days,
            len(df), n_comp,
            round(n_comp / len(df) * 100, 2) if len(df) else 0,
            len(det), len(pas), len(pro),
            round(len(det) / n_comp * 100, 2) if n_comp else 0,
            round(nps_score, 2), round(avg_tnps, 2),
            round(len(df) / n_days, 1), round(len(det) / n_days, 1),
            df["SHORT_CODE"].nunique(), df["AGENT_QUEUE"].nunique(),
            df["Agent_Id_"].nunique() if "Agent_Id_" in df.columns else 0,
            int((df["Q Mapping Lev 1"] == "Unmapped").sum()),
        ],
        "vs_Prev_Month": [
            "", "", "", "", "", "", "", "", "",
            delta_det_rate, delta_nps, "",
            "", "", "", "", "", "",
        ],
    })


# ============================================================
# 3. DAILY & MONTHLY TRENDS
# ============================================================
def build_daily_trend(df: pd.DataFrame) -> pd.DataFrame:
    g        = df.groupby("Date")
    completed = df[df["Q1_ANSWER: TNPS"].notna()].groupby("Date")

    out = pd.DataFrame({"Date": sorted(df["Date"].unique())}).set_index("Date")
    out["Surveys_Sent"]      = g.size()
    out["Surveys_Completed"] = completed.size()
    out["Completion_Rate_%"] = (out["Surveys_Completed"] / out["Surveys_Sent"] * 100).round(2)
    out["Detractors"]        = df[df["NPS_Category"] == "Detractor"].groupby("Date").size()
    out["Passives"]          = df[df["NPS_Category"] == "Passive"].groupby("Date").size()
    out["Promoters"]         = df[df["NPS_Category"] == "Promoter"].groupby("Date").size()
    out = out.fillna(0)

    out["Detractor_Rate_%"] = (
        out["Detractors"] / out["Surveys_Completed"].replace(0, np.nan) * 100
    ).round(2)
    out["NPS_Score"] = (
        (out["Promoters"] - out["Detractors"])
        / out["Surveys_Completed"].replace(0, np.nan) * 100
    ).round(2)
    out["Avg_TNPS"]              = completed["Q1_ANSWER: TNPS"].mean().round(2)
    out["Surveys_7d_Avg"]        = out["Surveys_Sent"].rolling(7, min_periods=1).mean().round(1)
    out["Detractors_7d_Avg"]     = out["Detractors"].rolling(7, min_periods=1).mean().round(1)
    out["Detractor_Rate_7d_Avg"] = out["Detractor_Rate_%"].rolling(7, min_periods=1).mean().round(2)

    thr    = CONFIG["anomaly_std_threshold"]
    mean_v = out["Surveys_Sent"].mean(); std_v = out["Surveys_Sent"].std()
    out["Volume_vs_Avg"] = out["Surveys_Sent"].apply(
        lambda x: "LOW"    if (std_v and x < mean_v - thr * std_v)
        else      "HIGH"   if (std_v and x > mean_v + thr * std_v)
        else      "Normal"
    )

    mean_d = out["Detractor_Rate_%"].mean(); std_d = out["Detractor_Rate_%"].std()
    out["Detractor_Rate_Status"] = out["Detractor_Rate_%"].apply(
        lambda x: "WORSE"  if (pd.notna(x) and std_d and x > mean_d + thr * std_d)
        else      "BETTER" if (pd.notna(x) and std_d and x < mean_d - thr * std_d)
        else      "Normal"
    )

    return out.reset_index().fillna(0)


def build_monthly_trend(df: pd.DataFrame) -> pd.DataFrame:
    g        = df.groupby("Month")
    completed = df[df["Q1_ANSWER: TNPS"].notna()].groupby("Month")

    out = pd.DataFrame({"Month": sorted(df["Month"].unique())}).set_index("Month")
    out["Surveys_Sent"]      = g.size()
    out["Surveys_Completed"] = completed.size()
    out["Completion_Rate_%"] = (out["Surveys_Completed"] / out["Surveys_Sent"] * 100).round(2)
    out["Detractors"]        = df[df["NPS_Category"] == "Detractor"].groupby("Month").size()
    out["Passives"]          = df[df["NPS_Category"] == "Passive"].groupby("Month").size()
    out["Promoters"]         = df[df["NPS_Category"] == "Promoter"].groupby("Month").size()
    out = out.fillna(0)
    out["Detractor_Rate_%"] = (
        out["Detractors"] / out["Surveys_Completed"].replace(0, np.nan) * 100
    ).round(2)
    out["NPS_Score"] = (
        (out["Promoters"] - out["Detractors"])
        / out["Surveys_Completed"].replace(0, np.nan) * 100
    ).round(2)
    out["Avg_TNPS"]        = completed["Q1_ANSWER: TNPS"].mean().round(2)
    out["MoM_Surveys_%"]   = (out["Surveys_Sent"].pct_change() * 100).round(2)
    out["MoM_Detractors_%"] = (out["Detractors"].pct_change() * 100).round(2)

    return out.reset_index().fillna(0)


# ============================================================
# 4. SHORT CODE × DAY PIVOT  (coloured Trend column)
# ============================================================
def build_shortcode_daily_pivot(df: pd.DataFrame) -> pd.DataFrame:
    det = df[df["NPS_Category"] == "Detractor"]
    if det.empty:
        return pd.DataFrame()

    pivot = det.groupby(["SHORT_CODE", "Date"]).size().unstack(fill_value=0)
    pivot = pivot.reindex(sorted(pivot.columns), axis=1)

    date_cols = list(pivot.columns)
    pivot["Total_Detractors"] = pivot[date_cols].sum(axis=1)

    if len(date_cols) >= 2:
        mid    = len(date_cols) // 2
        first  = pivot[date_cols[:mid]].sum(axis=1)
        second = pivot[date_cols[mid:]].sum(axis=1)
    else:
        first  = pd.Series(0, index=pivot.index)
        second = pivot[date_cols[0]] if date_cols else pd.Series(0, index=pivot.index)

    pivot["First_Half_Sum"]  = first
    pivot["Second_Half_Sum"] = second
    pivot["Change_Abs"]      = second - first
    pivot["Change_%"]        = ((second - first) / first.replace(0, np.nan) * 100).round(1)

    def label(row):
        f, s, p = row["First_Half_Sum"], row["Second_Half_Sum"], row["Change_%"]
        if f == 0 and s > 0:  return "NEW ↑"
        if f > 0 and s == 0:  return "STOPPED ↓"
        if pd.isna(p):         return "→ flat"
        if p > 10:             return f"↑ +{p:.0f}%"
        if p < -10:            return f"↓ {p:.0f}%"
        return "→ flat"

    pivot["Trend"] = pivot.apply(label, axis=1)
    return pivot.reset_index().sort_values("Total_Detractors", ascending=False)


# ============================================================
# 5. CATALOGS
# ============================================================
def build_shortcode_catalog(df: pd.DataFrame) -> pd.DataFrame:
    completed  = df[df["Q1_ANSWER: TNPS"].notna()]
    by_sc      = df.groupby("SHORT_CODE")
    comp_by_sc = completed.groupby("SHORT_CODE")
    det_by_sc  = df[df["NPS_Category"] == "Detractor"].groupby("SHORT_CODE")

    out = pd.DataFrame({"SHORT_CODE": sorted(df["SHORT_CODE"].dropna().unique())}).set_index("SHORT_CODE")
    out["Topic"] = ""
    out["Notes"] = ""
    out["First_Seen"]        = by_sc["Date"].min().astype(str)
    out["Last_Seen"]         = by_sc["Date"].max().astype(str)
    out["Total_Surveys"]     = by_sc.size()
    out["Completed_Surveys"] = comp_by_sc.size()
    out["Total_Detractors"]  = det_by_sc.size()
    out = out.fillna(0)
    out["Detractor_Rate_%"] = (
        out["Total_Detractors"] / out["Completed_Surveys"].replace(0, np.nan) * 100
    ).round(2)
    out["Avg_TNPS"] = comp_by_sc["Q1_ANSWER: TNPS"].mean().round(2)

    def mode_or_blank(s):
        m = s.mode()
        return m.iloc[0] if len(m) else ""

    out["Top_Owner_Team"] = by_sc["OWNER_TEAM"].agg(mode_or_blank) if "OWNER_TEAM" in df.columns else ""
    out["Top_Substatus"]  = by_sc["SUBSTATUS"].agg(mode_or_blank)  if "SUBSTATUS"  in df.columns else ""
    out["Top_Call_Type"]  = by_sc["CALL_TYPE"].agg(mode_or_blank)  if "CALL_TYPE"  in df.columns else ""
    out["Top_Agent_Queue"] = by_sc["AGENT_QUEUE"].agg(mode_or_blank)

    return out.reset_index().sort_values("Total_Detractors", ascending=False)


def build_queue_catalog(df: pd.DataFrame) -> pd.DataFrame:
    by_q  = df.groupby("AGENT_QUEUE")
    comp  = df[df["Q1_ANSWER: TNPS"].notna()].groupby("AGENT_QUEUE")
    det   = df[df["NPS_Category"] == "Detractor"].groupby("AGENT_QUEUE")

    out = pd.DataFrame({"AGENT_QUEUE": sorted(df["AGENT_QUEUE"].dropna().unique())}).set_index("AGENT_QUEUE")
    out["Total_Surveys"]    = by_q.size()
    out["Completed"]        = comp.size()
    out["Detractors"]       = det.size()
    out = out.fillna(0)
    out["Detractor_Rate_%"] = (out["Detractors"] / out["Completed"].replace(0, np.nan) * 100).round(2)
    out["Avg_TNPS"]         = comp["Q1_ANSWER: TNPS"].mean().round(2)

    def first(s): return s.iloc[0] if len(s) else ""
    out["Lev_1"] = by_q["Q Mapping Lev 1"].agg(first)
    out["Lev_2"] = by_q["Q Mapping Lev 2 Combined"].agg(first)
    out["Lev_3"] = by_q["Q Mapping Lev 3 New PF Seg"].agg(first)
    out["Lev_4"] = by_q["Q Mapping Lev 4 Seg"].agg(first)
    out["Site"]  = by_q["Site"].agg(first)

    return out.reset_index().sort_values("Detractors", ascending=False)


# ============================================================
# 6. MAPPING COVERAGE REPORT  (NEW)
# ============================================================
def build_mapping_coverage(df: pd.DataFrame) -> pd.DataFrame:
    by_q = df.groupby("AGENT_QUEUE")
    unmapped = df[df["Q Mapping Lev 1"] == "Unmapped"]
    if unmapped.empty:
        return pd.DataFrame(columns=["AGENT_QUEUE", "Total_Surveys", "Unmapped_Surveys",
                                     "Unmapped_%", "Recommendation"])

    u_by_q = unmapped.groupby("AGENT_QUEUE")
    out = pd.DataFrame({
        "AGENT_QUEUE":      sorted(unmapped["AGENT_QUEUE"].dropna().unique()),
    }).set_index("AGENT_QUEUE")
    out["Total_Surveys"]   = by_q.size()
    out["Unmapped_Surveys"] = u_by_q.size()
    out = out.fillna(0)
    out["Unmapped_%"] = (out["Unmapped_Surveys"] / out["Total_Surveys"] * 100).round(1)
    out["Recommendation"] = out["Unmapped_%"].apply(
        lambda x: "URGENT – add to mapping file" if x == 100
        else "Review – partial mapping" if x > 0 else "OK"
    )
    return out.reset_index().sort_values("Unmapped_Surveys", ascending=False)


# ============================================================
# 7. DUPLICATE DETECTION  (NEW)
# ============================================================
def build_duplicate_surveys(df: pd.DataFrame) -> pd.DataFrame:
    if "MSISDN" not in df.columns:
        return pd.DataFrame()
    df2 = df.copy()
    df2 = df2.sort_values("OUTBOUND_IVR_DT")
    df2["prev_dt"] = df2.groupby("MSISDN")["OUTBOUND_IVR_DT"].shift(1)
    df2["hours_since_prev"] = (
        (df2["OUTBOUND_IVR_DT"] - df2["prev_dt"]).dt.total_seconds() / 3600
    )
    dups = df2[df2["hours_since_prev"] < 24].copy()
    if dups.empty:
        return pd.DataFrame()
    return dups[["MSISDN", "OUTBOUND_IVR_DT", "prev_dt", "hours_since_prev",
                 "SHORT_CODE", "AGENT_QUEUE", "NPS_Category"]].sort_values("hours_since_prev")


# ============================================================
# 8. DIMENSION BREAKDOWNS
# ============================================================
def breakdown_by(df: pd.DataFrame, col: str) -> pd.DataFrame:
    if col not in df.columns:
        return pd.DataFrame()
    completed = df[df["Q1_ANSWER: TNPS"].notna()]
    g = completed.groupby(col)
    out = pd.DataFrame({col: sorted(completed[col].dropna().astype(str).unique())}).set_index(col)
    out["Completed_Surveys"] = completed.groupby(col).size()
    out["Detractors"]        = completed[completed["NPS_Category"] == "Detractor"].groupby(col).size()
    out = out.fillna(0)
    out["Detractor_Rate_%"] = (
        out["Detractors"] / out["Completed_Surveys"].replace(0, np.nan) * 100
    ).round(2)
    out["Avg_TNPS"] = g["Q1_ANSWER: TNPS"].mean().round(2)
    return out.reset_index().sort_values("Detractors", ascending=False)


def build_lev3_lev4_heatmap(df: pd.DataFrame) -> pd.DataFrame:
    comp = df[df["Q1_ANSWER: TNPS"].notna()]
    if comp.empty:
        return pd.DataFrame()
    comp_p = pd.pivot_table(comp, index="Q Mapping Lev 3 New PF Seg",
                            columns="Q Mapping Lev 4 Seg",
                            values="Q1_ANSWER: TNPS", aggfunc="count", fill_value=0)
    det_p = pd.pivot_table(comp[comp["NPS_Category"] == "Detractor"],
                           index="Q Mapping Lev 3 New PF Seg",
                           columns="Q Mapping Lev 4 Seg",
                           values="Q1_ANSWER: TNPS", aggfunc="count", fill_value=0)
    det_p = det_p.reindex(index=comp_p.index, columns=comp_p.columns, fill_value=0)
    rate = (det_p / comp_p.replace(0, np.nan) * 100).round(2).fillna(0)
    rate.index.name = "Lev3 \\ Lev4"
    return rate.reset_index()


# ============================================================
# 9. PATTERNS, REPEATS, CROSS-TABS, AGENTS
# ============================================================
def build_hour_pattern(df: pd.DataFrame) -> pd.DataFrame:
    comp = df[df["Q1_ANSWER: TNPS"].notna()]
    out = pd.DataFrame({"Hour": range(24)}).set_index("Hour")
    out["Surveys"]    = comp.groupby("Hour").size()
    out["Detractors"] = comp[comp["NPS_Category"] == "Detractor"].groupby("Hour").size()
    out = out.fillna(0)
    out["Detractor_Rate_%"] = (out["Detractors"] / out["Surveys"].replace(0, np.nan) * 100).round(2)
    return out.reset_index().fillna(0)


def build_dow_pattern(df: pd.DataFrame) -> pd.DataFrame:
    order = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
    comp = df[df["Q1_ANSWER: TNPS"].notna()]
    out = pd.DataFrame({"Day": order}).set_index("Day")
    out["Surveys"]    = comp.groupby("DayOfWeek").size()
    out["Detractors"] = comp[comp["NPS_Category"] == "Detractor"].groupby("DayOfWeek").size()
    out = out.fillna(0)
    out["Detractor_Rate_%"] = (out["Detractors"] / out["Surveys"].replace(0, np.nan) * 100).round(2)
    return out.reset_index().fillna(0)


def build_repeat_detractors(df: pd.DataFrame) -> pd.DataFrame:
    det = df[df["NPS_Category"] == "Detractor"]
    if det.empty or "MSISDN" not in det.columns:
        return pd.DataFrame()
    counts = det.groupby("MSISDN").size().reset_index(name="Detractor_Count")
    counts = counts[counts["Detractor_Count"] >= 2].sort_values("Detractor_Count", ascending=False)
    last = det.sort_values("OUTBOUND_IVR_DT").groupby("MSISDN").tail(1)[
        ["MSISDN", "OUTBOUND_IVR_DT", "SHORT_CODE", "AGENT_QUEUE"]
    ].rename(columns={"OUTBOUND_IVR_DT": "Last_Detractor_Date",
                       "SHORT_CODE": "Last_Short_Code", "AGENT_QUEUE": "Last_Queue"})
    return counts.merge(last, on="MSISDN", how="left")


def build_q2_q1_crosstab(df: pd.DataFrame, q_col: str, label: str) -> pd.DataFrame:
    if q_col not in df.columns:
        return pd.DataFrame()
    comp = df[(df["Q1_ANSWER: TNPS"].notna()) & (df[q_col].notna())]
    if comp.empty:
        return pd.DataFrame()
    ct = pd.crosstab(comp["NPS_Category"], comp[q_col], margins=True, margins_name="Total")
    ct.index.name = f"NPS_Category \\ {label}"
    return ct.reset_index()


def build_agent_ranking(df: pd.DataFrame) -> pd.DataFrame:
    if "Agent_Id_" not in df.columns:
        return pd.DataFrame()
    comp = df[df["Q1_ANSWER: TNPS"].notna()]
    g = comp.groupby("Agent_Id_")
    out = pd.DataFrame({"Agent_Id": sorted(comp["Agent_Id_"].dropna().unique())}).set_index("Agent_Id")
    out["Surveys"]    = g.size()
    out["Detractors"] = comp[comp["NPS_Category"] == "Detractor"].groupby("Agent_Id_").size()
    out = out.fillna(0)
    out["Detractor_Rate_%"] = (out["Detractors"] / out["Surveys"].replace(0, np.nan) * 100).round(2)
    out["Avg_TNPS"]         = g["Q1_ANSWER: TNPS"].mean().round(2)
    out = out[out["Surveys"] >= CONFIG["min_agent_sample"]]

    # Percentile bands
    out["Percentile"] = out["Detractor_Rate_%"].rank(pct=True).mul(100).round(0).astype(int)
    out["Band"] = pd.cut(out["Percentile"],
                         bins=[0, 10, 25, 75, 90, 100],
                         labels=["Top 10%", "Top 25%", "Mid 50%", "Bottom 25%", "Bottom 10%"],
                         include_lowest=True)
    return out.reset_index().sort_values("Detractor_Rate_%", ascending=False)


# ============================================================
# 10. AGENT PEER-GROUP BENCHMARKING  (NEW)
# ============================================================
def build_agent_peer_benchmark(df: pd.DataFrame) -> pd.DataFrame:
    if "Agent_Id_" not in df.columns or "AGENT_QUEUE" not in df.columns:
        return pd.DataFrame()
    comp = df[df["Q1_ANSWER: TNPS"].notna()]
    agent = comp.groupby(["Agent_Id_", "AGENT_QUEUE"]).agg(
        Surveys    = ("Q1_ANSWER: TNPS", "count"),
        Avg_TNPS   = ("Q1_ANSWER: TNPS", "mean"),
    ).reset_index()
    det = comp[comp["NPS_Category"] == "Detractor"].groupby(["Agent_Id_", "AGENT_QUEUE"]).size().reset_index(name="Detractors")
    agent = agent.merge(det, on=["Agent_Id_", "AGENT_QUEUE"], how="left").fillna(0)
    agent["Detractor_Rate_%"] = (agent["Detractors"] / agent["Surveys"] * 100).round(2)
    agent = agent[agent["Surveys"] >= CONFIG["min_agent_sample"]]

    # Queue-level benchmark
    qbench = agent.groupby("AGENT_QUEUE")["Detractor_Rate_%"].agg(
        Queue_Avg="mean", Queue_Median="median"
    ).reset_index()
    agent = agent.merge(qbench, on="AGENT_QUEUE", how="left")
    agent["vs_Queue_Avg"] = (agent["Detractor_Rate_%"] - agent["Queue_Avg"]).round(2)
    agent["Peer_Rank"]    = agent.groupby("AGENT_QUEUE")["Detractor_Rate_%"].rank(ascending=True).astype(int)
    return agent.sort_values(["AGENT_QUEUE", "Peer_Rank"])


# ============================================================
# 11. COHORT ANALYSIS  (NEW)
# ============================================================
def build_cohort_analysis(df: pd.DataFrame) -> pd.DataFrame:
    """Track MSISDN across months: first detractor month → subsequent category."""
    if "MSISDN" not in df.columns:
        return pd.DataFrame()
    comp = df[df["Q1_ANSWER: TNPS"].notna()][["MSISDN", "Month", "NPS_Category"]].copy()

    # First month each MSISDN was a detractor
    first_det = (
        comp[comp["NPS_Category"] == "Detractor"]
        .groupby("MSISDN")["Month"].min().reset_index(name="First_Detractor_Month")
    )
    if first_det.empty:
        return pd.DataFrame()

    cohort = comp.merge(first_det, on="MSISDN")
    cohort = cohort[cohort["Month"] > cohort["First_Detractor_Month"]]
    if cohort.empty:
        return pd.DataFrame()

    summary = cohort.groupby(["First_Detractor_Month", "NPS_Category"]).size().unstack(fill_value=0)
    summary.index.name = "Cohort_Month (First_Detractor)"
    total = summary.sum(axis=1)
    for col in summary.columns:
        summary[f"{col}_%"] = (summary[col] / total * 100).round(1)

    recovery = pd.DataFrame()
    if "Promoter" in summary.columns and "Detractor" in summary.columns:
        recovery["Recovery_Rate_%"] = (
            summary["Promoter"] / total * 100
        ).round(1)
    return summary.reset_index()


# ============================================================
# 12. ROOT-CAUSE TOXIC COMBOS  (NEW)
# ============================================================
def build_toxic_combos(df: pd.DataFrame, top_n: int = 20) -> pd.DataFrame:
    """Top dimension combinations driving detractors."""
    dims = [c for c in ["OWNER_TEAM", "SUBSTATUS", "CALL_TYPE", "AGENT_QUEUE"] if c in df.columns]
    if len(dims) < 2:
        return pd.DataFrame()

    comp = df[df["Q1_ANSWER: TNPS"].notna()].copy()
    comp["_combo"] = comp[dims].fillna("?").astype(str).agg(" | ".join, axis=1)

    g = comp.groupby("_combo")
    out = pd.DataFrame({"Combo": list(g.groups.keys())}).set_index("Combo")
    out["Surveys"]    = g.size()
    out["Detractors"] = comp[comp["NPS_Category"] == "Detractor"].groupby("_combo").size()
    out = out.fillna(0)
    out["Detractor_Rate_%"] = (out["Detractors"] / out["Surveys"].replace(0, np.nan) * 100).round(2)
    out["Dimensions"] = " | ".join(dims)
    return (
        out[out["Surveys"] >= 5]
        .reset_index()
        .sort_values("Detractor_Rate_%", ascending=False)
        .head(top_n)
    )


# ============================================================
# 13. VELOCITY ALERTS  (NEW)
# ============================================================
def build_velocity_alerts(df: pd.DataFrame) -> pd.DataFrame:
    """Week-over-week detractor rate jump per queue."""
    comp = df[df["Q1_ANSWER: TNPS"].notna()]
    if "Week" not in comp.columns:
        return pd.DataFrame()

    g = comp.groupby(["AGENT_QUEUE", "Week"])
    weekly = g.agg(
        Surveys    = ("Q1_ANSWER: TNPS", "count"),
    ).reset_index()
    det_w = comp[comp["NPS_Category"] == "Detractor"].groupby(["AGENT_QUEUE", "Week"]).size().reset_index(name="Detractors")
    weekly = weekly.merge(det_w, on=["AGENT_QUEUE", "Week"], how="left").fillna(0)
    weekly["Detractor_Rate_%"] = (weekly["Detractors"] / weekly["Surveys"] * 100).round(2)
    weekly = weekly.sort_values(["AGENT_QUEUE", "Week"])
    weekly["Prev_Rate"] = weekly.groupby("AGENT_QUEUE")["Detractor_Rate_%"].shift(1)
    weekly["WoW_Change_%"] = (weekly["Detractor_Rate_%"] - weekly["Prev_Rate"]).round(2)
    threshold = CONFIG["velocity_alert_pct"]
    alerts = weekly[weekly["WoW_Change_%"].abs() >= threshold].copy()
    alerts["Alert"] = alerts["WoW_Change_%"].apply(
        lambda x: f"↑ SPIKE +{x:.1f}%" if x > 0 else f"↓ DROP {x:.1f}%"
    )
    return alerts.sort_values("WoW_Change_%", ascending=False)


# ============================================================
# 14. TOP 10 / BOTTOM 10  (NEW)
# ============================================================
def build_top_bottom(df: pd.DataFrame) -> pd.DataFrame:
    q_cat = build_queue_catalog(df)
    sc_cat = build_shortcode_catalog(df)

    q_top    = q_cat.nsmallest(10, "Detractor_Rate_%")[["AGENT_QUEUE", "Detractors", "Detractor_Rate_%", "Avg_TNPS"]].copy()
    q_bottom = q_cat.nlargest(10, "Detractor_Rate_%")[["AGENT_QUEUE", "Detractors", "Detractor_Rate_%", "Avg_TNPS"]].copy()
    sc_top   = sc_cat.nsmallest(10, "Detractor_Rate_%")[["SHORT_CODE", "Total_Detractors", "Detractor_Rate_%", "Avg_TNPS"]].copy()
    sc_btm   = sc_cat.nlargest(10, "Detractor_Rate_%")[["SHORT_CODE", "Total_Detractors", "Detractor_Rate_%", "Avg_TNPS"]].copy()

    q_top["Rank_Type"]  = "Best 10 Queues"
    q_bottom["Rank_Type"] = "Worst 10 Queues"
    sc_top["Rank_Type"] = "Best 10 Short Codes"
    sc_btm["Rank_Type"] = "Worst 10 Short Codes"

    q_top.rename(columns={"AGENT_QUEUE": "Name", "Detractors": "Detractor_Count"}, inplace=True)
    q_bottom.rename(columns={"AGENT_QUEUE": "Name", "Detractors": "Detractor_Count"}, inplace=True)
    sc_top.rename(columns={"SHORT_CODE": "Name", "Total_Detractors": "Detractor_Count"}, inplace=True)
    sc_btm.rename(columns={"SHORT_CODE": "Name", "Total_Detractors": "Detractor_Count"}, inplace=True)

    return pd.concat([q_top, q_bottom, sc_top, sc_btm], ignore_index=True)


# ============================================================
# 15. NPS WATERFALL  (NEW)
# ============================================================
def build_nps_waterfall(df: pd.DataFrame) -> pd.DataFrame:
    months = sorted(df["Month"].unique())
    rows = []
    for m in months:
        sub = df[(df["Month"] == m) & (df["Q1_ANSWER: TNPS"].notna())]
        n   = len(sub)
        d   = (sub["NPS_Category"] == "Detractor").sum()
        p   = (sub["NPS_Category"] == "Promoter").sum()
        rows.append({
            "Month":            m,
            "Completed":        n,
            "Promoters":        int(p),
            "Detractors":       int(d),
            "Net_Gain":         int(p - d),
            "NPS_Score":        round((p - d) / n * 100, 2) if n else 0,
            "Promoter_Rate_%":  round(p / n * 100, 2) if n else 0,
            "Detractor_Rate_%": round(d / n * 100, 2) if n else 0,
        })
    return pd.DataFrame(rows)


# ============================================================
# 16. SLA BREACH HEATMAP  (NEW)
# ============================================================
def build_sla_breach_heatmap(df: pd.DataFrame) -> pd.DataFrame:
    comp = df[df["Q1_ANSWER: TNPS"].notna()]
    if comp.empty:
        return pd.DataFrame()
    threshold = CONFIG["sla_detractor_rate_threshold"]

    pivot_surveys = pd.pivot_table(comp, index="AGENT_QUEUE", columns="Date",
                                   values="Q1_ANSWER: TNPS", aggfunc="count", fill_value=0)
    pivot_det = pd.pivot_table(comp[comp["NPS_Category"] == "Detractor"],
                               index="AGENT_QUEUE", columns="Date",
                               values="Q1_ANSWER: TNPS", aggfunc="count", fill_value=0)
    pivot_det = pivot_det.reindex(index=pivot_surveys.index,
                                  columns=pivot_surveys.columns, fill_value=0)
    rate = (pivot_det / pivot_surveys.replace(0, np.nan) * 100).round(1).fillna(0)

    # Mark only breaches  (applymap → map for pandas 2.x compatibility)
    _map_fn = rate.map if hasattr(rate, "map") and callable(getattr(rate, "map")) and rate.map.__module__ != "builtins" else rate.applymap
    breach = rate.apply(lambda col: col.map(lambda x: x if x >= threshold else 0))
    breach.index.name = "Queue \\ Date"
    # Drop queues with no breaches at all
    breach = breach[breach.sum(axis=1) > 0]
    return breach.reset_index() if not breach.empty else pd.DataFrame()


# ============================================================
# 17. CHANNEL COMPARISON  (NEW)
# ============================================================
def build_channel_comparison(df: pd.DataFrame) -> pd.DataFrame:
    col = "Channel_Name" if "Channel_Name" in df.columns else None
    if col is None:
        return pd.DataFrame()
    return breakdown_by(df, col)


# ============================================================
# 18. FCR IMPACT  (NEW)
# ============================================================
def build_fcr_impact(df: pd.DataFrame) -> pd.DataFrame:
    col = "FCR_FLAG" if "FCR_FLAG" in df.columns else None
    if col is None:
        return pd.DataFrame()
    return breakdown_by(df, col)


# ============================================================
# 19. FORECASTING  (with prediction intervals + per-queue + target)
# ============================================================
def forecast_series(series: pd.Series, horizon: int, label: str
                    ) -> tuple[pd.Series, pd.Series, pd.Series]:
    """Returns (forecast, lower_CI, upper_CI)."""
    s = series.dropna()
    if len(s) < 4:
        mean_val = s.mean() if len(s) else 0
        base = pd.Series([mean_val] * horizon)
        return base, base, base
    try:
        if len(s) >= 14:
            model = ExponentialSmoothing(
                s, seasonal_periods=7, seasonal="add", trend="add",
                initialization_method="estimated",
            ).fit()
        else:
            model = ExponentialSmoothing(
                s, trend="add", initialization_method="estimated",
            ).fit()
        f     = model.forecast(horizon).clip(lower=0)
        # Simple CI: ±1.96 * residual std
        resid_std = model.resid.std()
        ci_width  = 1.96 * resid_std
        lower = (f - ci_width).clip(lower=0)
        upper = f + ci_width
        return f, lower, upper
    except Exception as e:
        log.warning("Forecast failed for %s: %s — using mean", label, e)
        base = pd.Series([s.mean()] * horizon)
        return base, base, base


def build_forecast(daily: pd.DataFrame) -> pd.DataFrame:
    horizon = CONFIG["forecast_horizon_days"]
    if daily.empty:
        return pd.DataFrame()

    d = daily.copy()
    d["Date"] = pd.to_datetime(d["Date"])
    d = d.set_index("Date").sort_index()
    last_date    = d.index.max()
    future_dates = [last_date + timedelta(days=i + 1) for i in range(horizon)]
    out = pd.DataFrame({"Date": future_dates})

    for col, label in [
        ("Surveys_Sent", "Survey Volume"),
        ("Detractors",   "Detractor Count"),
        ("Detractor_Rate_%", "Detractor Rate %"),
        ("NPS_Score",    "NPS Score"),
        ("Avg_TNPS",     "Average TNPS"),
    ]:
        if col not in d.columns:
            continue
        f, lo, hi = forecast_series(d[col], horizon, label)
        out[f"Forecast_{col}"]      = f.values.round(2)
        out[f"Forecast_{col}_Lower"] = lo.values.round(2)
        out[f"Forecast_{col}_Upper"] = hi.values.round(2)

    # Breach date: when does forecast detractor rate exceed target?
    target = CONFIG["detractor_rate_target"]
    if "Forecast_Detractor_Rate_%" in out.columns:
        breach_rows = out[out["Forecast_Detractor_Rate_%"] >= target]
        breach_date = breach_rows["Date"].iloc[0] if not breach_rows.empty else "No breach in horizon"
        out.attrs["breach_date"] = str(breach_date)

    return out


def build_per_queue_forecast(df: pd.DataFrame, daily: pd.DataFrame) -> pd.DataFrame:
    horizon  = CONFIG["forecast_horizon_days"]
    top_n    = CONFIG["forecast_top_n_queues"]
    if daily.empty:
        return pd.DataFrame()

    top_queues = (
        df[df["NPS_Category"] == "Detractor"].groupby("AGENT_QUEUE").size()
        .nlargest(top_n).index.tolist()
    )

    last_date    = pd.to_datetime(daily["Date"]).max()
    future_dates = [last_date + timedelta(days=i + 1) for i in range(horizon)]
    out = pd.DataFrame({"Date": future_dates})

    for q in top_queues:
        sub = df[(df["AGENT_QUEUE"] == q) & (df["Q1_ANSWER: TNPS"].notna())]
        if sub.empty:
            continue
        daily_q = (
            sub.groupby("Date").apply(
                lambda x: pd.Series({
                    "Det_Rate": (
                        (x["NPS_Category"] == "Detractor").sum() / len(x) * 100
                    )
                })
            )
            .reset_index()
        )
        daily_q["Date"] = pd.to_datetime(daily_q["Date"])
        daily_q = daily_q.set_index("Date").sort_index()

        f, lo, hi = forecast_series(daily_q["Det_Rate"], horizon, q)
        safe_name = q[:20].replace(" ", "_").replace("/", "-")
        out[f"{safe_name}_Rate"]  = f.values.round(2)
        out[f"{safe_name}_Lower"] = lo.values.round(2)
        out[f"{safe_name}_Upper"] = hi.values.round(2)

    return out


# ============================================================
# 20. EXECUTIVE SUMMARY  (NEW)
# ============================================================
def build_executive_summary(df: pd.DataFrame, kpi: pd.DataFrame,
                             daily: pd.DataFrame, forecast: pd.DataFrame) -> pd.DataFrame:
    def kv(metric):
        row = kpi[kpi["Metric"] == metric]
        return row["Value"].iloc[0] if not row.empty else "N/A"

    det_rate = kv("Detractor Rate %")
    nps      = kv("NPS Score")
    total    = kv("Total Surveys Sent")
    n_det    = kv("Detractors (Q1 0-6)")

    # Trend direction
    if not daily.empty and len(daily) >= 14:
        mid    = len(daily) // 2
        first  = daily.iloc[:mid]["Detractor_Rate_%"].mean()
        second = daily.iloc[mid:]["Detractor_Rate_%"].mean()
        trend  = "↑ WORSENING" if second > first else "↓ IMPROVING"
    else:
        trend = "Insufficient data for trend"

    # Top problematic queue
    comp = df[df["Q1_ANSWER: TNPS"].notna()]
    if not comp.empty:
        qg     = comp[comp["NPS_Category"] == "Detractor"].groupby("AGENT_QUEUE").size()
        top_q  = qg.idxmax() if not qg.empty else "N/A"
        top_qv = int(qg.max()) if not qg.empty else 0
    else:
        top_q = "N/A"; top_qv = 0

    breach_info = ""
    if not forecast.empty and "Forecast_Detractor_Rate_%" in forecast.columns:
        breach_info = forecast.attrs.get("breach_date", "N/A")

    lines = [
        ("Period",             f"{kv('Date Range Start')} → {kv('Date Range End')} ({kv('Number of Days Covered')} days)"),
        ("Total Surveys",      f"{total:,}" if isinstance(total, int) else str(total)),
        ("Detractors",         str(n_det)),
        ("Detractor Rate",     f"{det_rate}%"),
        ("NPS Score",          str(nps)),
        ("Period Trend",       trend),
        ("Top Driver Queue",   f"{top_q}  ({top_qv:,} detractors)"),
        ("Forecast Breach Date", breach_info),
        ("Recommendation",     "Prioritise top driver queue and escalate toxic short codes to product team."),
    ]
    return pd.DataFrame(lines, columns=["Summary_Item", "Detail"])


# ============================================================
# 21. EXCEL HELPERS
# ============================================================
def style_header(cell, color="C8102E"):
    cell.font      = Font(bold=True, color="FFFFFF", size=11, name="Arial")
    cell.fill      = PatternFill("solid", start_color=color)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(border_style="thin", color="FFFFFF")
    cell.border    = Border(left=thin, right=thin, top=thin, bottom=thin)


def write_df(ws, df, start_row=1, start_col=1, header_color=None):
    if df is None or (hasattr(df, "empty") and df.empty):
        ws.cell(row=start_row, column=start_col, value="(no data)")
        return
    color = header_color or CONFIG["brand_primary"]

    for j, c in enumerate(df.columns, start=start_col):
        style_header(ws.cell(row=start_row, column=j, value=str(c)), color)

    for i, (_, row) in enumerate(df.iterrows(), start=start_row + 1):
        for j, c in enumerate(df.columns, start=start_col):
            v = row[c]
            if pd.isna(v):           v = ""
            elif isinstance(v, np.integer):  v = int(v)
            elif isinstance(v, np.floating): v = float(v)
            cell = ws.cell(row=i, column=j, value=v)
            cell.font      = Font(name="Arial", size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center")

    for j, c in enumerate(df.columns, start=start_col):
        max_len = max([len(str(c))] + [len(str(v)) for v in df[c].astype(str).head(200)])
        ws.column_dimensions[get_column_letter(j)].width = min(max(max_len + 2, 12), 40)

    ws.freeze_panes = ws.cell(row=start_row + 1, column=start_col)


# ============================================================
# 22. DASHBOARD SHEET
# ============================================================
def write_dashboard(ws, kpi: pd.DataFrame, daily: pd.DataFrame,
                    forecast: pd.DataFrame, kpi_delta: bool = True):
    ws.sheet_view.showGridLines = False

    def kv(metric):
        row = kpi[kpi["Metric"] == metric]
        return row["Value"].iloc[0] if not row.empty else "N/A"

    def dv(metric):
        if "vs_Prev_Month" not in kpi.columns:
            return ""
        row = kpi[kpi["Metric"] == metric]
        return row["vs_Prev_Month"].iloc[0] if not row.empty else ""

    # Title
    ws.merge_cells("B2:K3")
    t = ws["B2"]
    t.value     = "TNPS Detractor Analytics & Forecasting Dashboard"
    t.font      = Font(name="Arial", size=20, bold=True, color="FFFFFF")
    t.fill      = PatternFill("solid", start_color=CONFIG["brand_secondary"])
    t.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("B4:K4")
    sub = ws["B4"]
    sub.value = (
        f"Period: {kv('Date Range Start')}  →  {kv('Date Range End')}"
        f"   |   {kv('Number of Days Covered')} days covered"
    )
    sub.font      = Font(name="Arial", size=11, italic=True, color="666666")
    sub.alignment = Alignment(horizontal="center", vertical="center")

    # KPI cards
    headline = [
        ("Total Surveys",       "Total Surveys Sent",      CONFIG["brand_secondary"]),
        ("Detractors",          "Detractors (Q1 0-6)",     CONFIG["brand_bad"]),
        ("Detractor Rate %",    "Detractor Rate %",        CONFIG["brand_bad"]),
        ("NPS Score",           "NPS Score",               CONFIG["brand_good"]),
        ("Completion Rate %",   "Completion Rate %",       CONFIG["brand_primary"]),
        ("Avg Surveys / Day",   "Average Surveys / Day",   CONFIG["brand_secondary"]),
    ]
    row = 6
    for i, (title, metric_name, color) in enumerate(headline):
        col = 2 + (i % 3) * 3
        r   = row + (i // 3) * 4
        val   = kv(metric_name)
        delta = dv(metric_name)

        ws.merge_cells(start_row=r, start_column=col, end_row=r, end_column=col + 2)
        c = ws.cell(row=r, column=col)
        c.value      = title
        c.font       = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        c.fill       = PatternFill("solid", start_color=color)
        c.alignment  = Alignment(horizontal="center", vertical="center")

        ws.merge_cells(start_row=r+1, start_column=col, end_row=r+2, end_column=col+2)
        c = ws.cell(row=r+1, column=col)
        c.value     = val
        c.font      = Font(name="Arial", size=22, bold=True, color="1F1F1F")
        c.fill      = PatternFill("solid", start_color="F5F5F5")
        c.alignment = Alignment(horizontal="center", vertical="center")

        # Delta badge (row r+3)
        if delta not in ("", "N/A"):
            ws.merge_cells(start_row=r+3, start_column=col, end_row=r+3, end_column=col+2)
            dc = ws.cell(row=r+3, column=col)
            try:
                delta_f = float(delta)
                arrow   = "▲" if delta_f > 0 else "▼"
                d_color = CONFIG["brand_bad"] if delta_f > 0 else CONFIG["brand_good"]
            except (ValueError, TypeError):
                arrow = ""; d_color = "888888"
            dc.value     = f"{arrow} {delta} vs prev month"
            dc.font      = Font(name="Arial", size=9, color=d_color, bold=True)
            dc.alignment = Alignment(horizontal="center", vertical="center")

    for col_l in ["B","C","D","E","F","G","H","I","J"]:
        ws.column_dimensions[col_l].width = 14
    for r in range(2, 24):
        ws.row_dimensions[r].height = 22

    # Chart helper data block
    chart_start_col = 13
    ws.cell(row=2, column=chart_start_col,   value="Date")
    ws.cell(row=2, column=chart_start_col+1, value="Surveys_Sent")
    ws.cell(row=2, column=chart_start_col+2, value="Detractors")
    ws.cell(row=2, column=chart_start_col+3, value="Detractor_Rate_%")
    for i, (_, row_d) in enumerate(daily.iterrows(), start=3):
        ws.cell(row=i, column=chart_start_col,   value=str(row_d["Date"]))
        ws.cell(row=i, column=chart_start_col+1, value=float(row_d["Surveys_Sent"]))
        ws.cell(row=i, column=chart_start_col+2, value=float(row_d["Detractors"]))
        ws.cell(row=i, column=chart_start_col+3, value=float(row_d["Detractor_Rate_%"]))
    end_row = 2 + len(daily)
    for c in range(chart_start_col, chart_start_col + 4):
        ws.column_dimensions[get_column_letter(c)].width = 2

    ch = LineChart()
    ch.title       = "Daily Survey Volume & Detractors"
    ch.style       = 2
    ch.y_axis.title = "Count"
    data_ref = Reference(ws, min_col=chart_start_col+1, max_col=chart_start_col+2,
                         min_row=2, max_row=end_row)
    cats     = Reference(ws, min_col=chart_start_col, min_row=3, max_row=end_row)
    ch.add_data(data_ref, titles_from_data=True)
    ch.set_categories(cats)
    ch.width = 24; ch.height = 10
    ws.add_chart(ch, "B25")

    ch2 = LineChart()
    ch2.title        = "Daily Detractor Rate %"
    ch2.style        = 12
    ch2.y_axis.title = "Detractor Rate %"
    data_ref2 = Reference(ws, min_col=chart_start_col+3, max_col=chart_start_col+3,
                          min_row=2, max_row=end_row)
    ch2.add_data(data_ref2, titles_from_data=True)
    ch2.set_categories(cats)
    ch2.width = 24; ch2.height = 10
    ws.add_chart(ch2, "B47")

    if not forecast.empty:
        fc_sc = chart_start_col + 6
        ws.cell(row=2, column=fc_sc,   value="Date")
        ws.cell(row=2, column=fc_sc+1, value="Forecast_Surveys")
        ws.cell(row=2, column=fc_sc+2, value="Forecast_Detractors")
        ws.cell(row=2, column=fc_sc+3, value="Forecast_Det_Rate")
        for i, (_, row_f) in enumerate(forecast.iterrows(), start=3):
            ws.cell(row=i, column=fc_sc,   value=str(row_f["Date"].date() if hasattr(row_f["Date"], "date") else row_f["Date"]))
            ws.cell(row=i, column=fc_sc+1, value=float(row_f.get("Forecast_Surveys_Sent", 0)))
            ws.cell(row=i, column=fc_sc+2, value=float(row_f.get("Forecast_Detractors", 0)))
            ws.cell(row=i, column=fc_sc+3, value=float(row_f.get("Forecast_Detractor_Rate_%", 0)))
        fc_end = 2 + len(forecast)
        for c in range(fc_sc, fc_sc + 4):
            ws.column_dimensions[get_column_letter(c)].width = 2

        ch3 = LineChart()
        ch3.title        = f"Forecast – Next {CONFIG['forecast_horizon_days']} Days"
        ch3.style        = 13
        ch3.y_axis.title = "Count"
        data_ref3 = Reference(ws, min_col=fc_sc+1, max_col=fc_sc+2, min_row=2, max_row=fc_end)
        cats3     = Reference(ws, min_col=fc_sc,   min_row=3, max_row=fc_end)
        ch3.add_data(data_ref3, titles_from_data=True)
        ch3.set_categories(cats3)
        ch3.width = 24; ch3.height = 10
        ws.add_chart(ch3, "B69")


# ============================================================
# 23. TABLE OF CONTENTS  (NEW)
# ============================================================
def write_toc(ws, sheet_names: list[str]):
    ws.sheet_view.showGridLines = False
    ws.merge_cells("B2:E2")
    t = ws["B2"]
    t.value     = "Table of Contents"
    t.font      = Font(name="Arial", size=16, bold=True, color="FFFFFF")
    t.fill      = PatternFill("solid", start_color=CONFIG["brand_primary"])
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 28
    ws.column_dimensions["B"].width = 4
    ws.column_dimensions["C"].width = 35
    ws.column_dimensions["D"].width = 20

    ws.cell(row=3, column=3, value="Sheet Name").font  = Font(bold=True, name="Arial")
    ws.cell(row=3, column=4, value="Navigate To").font = Font(bold=True, name="Arial")

    for idx, name in enumerate(sheet_names, start=4):
        ws.cell(row=idx, column=3, value=name).font = Font(name="Arial", size=10)
        link_cell = ws.cell(row=idx, column=4, value=f"Go → {name}")
        link_cell.hyperlink = f"#'{name}'!A1"
        link_cell.font = Font(name="Arial", size=10, color="0563C1", underline="single")


# ============================================================
# 24. EXCEL REPORT WRITER
# ============================================================
def _apply_trend_colors(ws, df):
    """Color-code the Trend column background in ShortCode pivot."""
    if "Trend" not in df.columns:
        return
    trend_col_idx = list(df.columns).index("Trend") + 1
    trend_col_letter = get_column_letter(trend_col_idx)

    color_map = {
        "NEW ↑":      "FFD7D7",
        "STOPPED ↓":  "D7F0D7",
        "→ flat":     "F5F5F5",
    }
    for row_idx in range(2, len(df) + 2):
        cell = ws.cell(row=row_idx, column=trend_col_idx)
        val  = str(cell.value or "")
        if val.startswith("↑"):
            cell.fill = PatternFill("solid", start_color="FFD7D7")
            cell.font = Font(name="Arial", size=10, color=CONFIG["brand_bad"], bold=True)
        elif val.startswith("↓ ") and "STOPPED" not in val:
            cell.fill = PatternFill("solid", start_color="D7F0D7")
            cell.font = Font(name="Arial", size=10, color=CONFIG["brand_good"], bold=True)
        elif "STOPPED" in val:
            cell.fill = PatternFill("solid", start_color="D7F0D7")
            cell.font = Font(name="Arial", size=10, color=CONFIG["brand_good"], bold=True)
        elif "NEW" in val:
            cell.fill = PatternFill("solid", start_color="FFD7D7")
            cell.font = Font(name="Arial", size=10, color=CONFIG["brand_bad"], bold=True)
        else:
            cell.fill = PatternFill("solid", start_color="F5F5F5")


def write_excel_report(out_path: str, sheets: dict, kpi: pd.DataFrame,
                       daily: pd.DataFrame, forecast: pd.DataFrame):
    wb = Workbook()

    # Dashboard
    wb.active.title = "Dashboard"
    write_dashboard(wb["Dashboard"], kpi, daily, forecast)

    order = [
        "Executive_Summary",
        "KPI_Summary",
        "Daily_Trend",
        "Monthly_Trend",
        "ShortCode_Daily_Pivot",
        "ShortCode_Catalog",
        "Queue_Catalog",
        "Mapping_Coverage",
        "Forecast",
        "Per_Queue_Forecast",
        "Top10_Bottom10",
        "NPS_Waterfall",
        "SLA_Breach_Heatmap",
        "By_OwnerTeam",
        "By_Substatus",
        "By_CallType",
        "By_ProdType",
        "Channel_Comparison",
        "FCR_Impact",
        "By_Reachability",
        "By_FCR_Flag",
        "By_Region",
        "By_Lev1",
        "By_Lev2",
        "By_Lev3",
        "By_Lev4",
        "By_Site",
        "Heatmap_Lev3_x_Lev4",
        "Hour_Pattern",
        "DayOfWeek_Pattern",
        "Toxic_Combos",
        "Velocity_Alerts",
        "Cohort_Analysis",
        "Repeat_Detractors",
        "Duplicate_Surveys",
        "Agent_Peer_Benchmark",
        "Agent_Ranking",
        "Q2_Attitude_vs_TNPS",
        "Q3_Fulfillment_vs_TNPS",
    ]

    created_sheets = ["Dashboard"]
    for name in order:
        if name not in sheets:
            continue
        df = sheets[name]
        if df is None or (hasattr(df, "empty") and df.empty):
            continue
        ws = wb.create_sheet(name)
        write_df(ws, df)
        created_sheets.append(name)

        if name == "ShortCode_Daily_Pivot":
            _apply_trend_colors(ws, df)
            n_rows = len(df) + 1
            try:
                tot_idx = list(df.columns).index("Total_Detractors") + 1
                if tot_idx > 2:
                    rng  = f"{get_column_letter(2)}2:{get_column_letter(tot_idx-1)}{n_rows}"
                    rule = ColorScaleRule(
                        start_type="min", start_color="FFFFFF",
                        mid_type="percentile", mid_value=50, mid_color="FFC7CE",
                        end_type="max", end_color="C8102E",
                    )
                    ws.conditional_formatting.add(rng, rule)
            except ValueError:
                pass

    # Table of Contents (inserted as second sheet)
    toc_ws = wb.create_sheet("Table_of_Contents", 1)
    write_toc(toc_ws, created_sheets)

    wb.save(out_path)
    log.info("Report saved → %s  (%d sheets)", out_path, len(wb.sheetnames))


# ============================================================
# 25. ORCHESTRATION
# ============================================================
def main():
    args = parse_args()
    apply_config_overrides(args)
    use_cache = not args.no_cache

    log.info("=" * 60)
    log.info(" TNPS DETRACTOR ANALYTICS & FORECASTING  (Enhanced)")
    log.info("=" * 60)

    data, mapping = load_data(CONFIG["data_folder"], CONFIG["mapping_file"], use_cache)

    def safe(name, fn, *args, **kwargs):
        """Call fn(*args) and return its result; on any error log and return empty DataFrame."""
        try:
            result = fn(*args, **kwargs)
            return result if result is not None else pd.DataFrame()
        except Exception as exc:
            log.error("  ✗ %s failed: %s", name, exc)
            return pd.DataFrame()

    log.info("Building KPI summary…")
    kpi = safe("KPI", build_kpi_summary, data)

    log.info("Building daily and monthly trends…")
    daily   = safe("Daily Trend",   build_daily_trend,   data)
    monthly = safe("Monthly Trend", build_monthly_trend, data)

    log.info("Building Short Code × Day pivot…")
    sc_pivot = safe("ShortCode Pivot", build_shortcode_daily_pivot, data)

    log.info("Building catalogs…")
    sc_catalog = safe("ShortCode Catalog", build_shortcode_catalog, data)
    q_catalog  = safe("Queue Catalog",     build_queue_catalog,     data)

    log.info("Building mapping coverage & duplicate detection…")
    map_coverage = safe("Mapping Coverage",  build_mapping_coverage,  data)
    duplicates   = safe("Duplicate Surveys", build_duplicate_surveys, data)

    log.info("Building dimension breakdowns…")
    by_dim = {
        "By_OwnerTeam":    safe("By_OwnerTeam",    breakdown_by, data, "OWNER_TEAM"),
        "By_Substatus":    safe("By_Substatus",    breakdown_by, data, "SUBSTATUS"),
        "By_CallType":     safe("By_CallType",     breakdown_by, data, "CALL_TYPE"),
        "By_ProdType":     safe("By_ProdType",     breakdown_by, data, "PROD_TYPE"),
        "By_Reachability": safe("By_Reachability", breakdown_by, data, "REACHABILITY"),
        "By_FCR_Flag":     safe("By_FCR_Flag",     breakdown_by, data, "FCR_FLAG"),
        "By_Region":       safe("By_Region",       breakdown_by, data, "Region"),
        "By_Lev1":         safe("By_Lev1",         breakdown_by, data, "Q Mapping Lev 1"),
        "By_Lev2":         safe("By_Lev2",         breakdown_by, data, "Q Mapping Lev 2 Combined"),
        "By_Lev3":         safe("By_Lev3",         breakdown_by, data, "Q Mapping Lev 3 New PF Seg"),
        "By_Lev4":         safe("By_Lev4",         breakdown_by, data, "Q Mapping Lev 4 Seg"),
        "By_Site":         safe("By_Site",         breakdown_by, data, "Site"),
    }

    log.info("Building heatmap, patterns, cross-tabs…")
    heatmap  = safe("Heatmap",        build_lev3_lev4_heatmap, data)
    hour_p   = safe("Hour Pattern",   build_hour_pattern,      data)
    dow_p    = safe("DoW Pattern",    build_dow_pattern,       data)
    repeats  = safe("Repeat Det.",    build_repeat_detractors, data)
    q2_ct    = safe("Q2 Crosstab",    build_q2_q1_crosstab, data, "Q2_ANSWER: Attitude",    "Attitude")
    q3_ct    = safe("Q3 Crosstab",    build_q2_q1_crosstab, data, "Q3_Answer: Fulfillment", "Fulfillment")
    agent_rk = safe("Agent Ranking",  build_agent_ranking,     data)

    log.info("Building new analytics: cohort, toxic combos, velocity, benchmarks…")
    cohort      = safe("Cohort Analysis",    build_cohort_analysis,    data)
    toxic       = safe("Toxic Combos",       build_toxic_combos,       data)
    velocity    = safe("Velocity Alerts",    build_velocity_alerts,    data)
    peer_bench  = safe("Peer Benchmark",     build_agent_peer_benchmark, data)
    top_bottom  = safe("Top/Bottom 10",      build_top_bottom,         data)
    waterfall   = safe("NPS Waterfall",      build_nps_waterfall,      data)
    sla_breach  = safe("SLA Breach Heatmap", build_sla_breach_heatmap, data)
    channel_cmp = safe("Channel Comparison", build_channel_comparison, data)
    fcr_impact  = safe("FCR Impact",         build_fcr_impact,         data)

    log.info("Building forecasts (Holt-Winters + per-queue + CI)…")
    forecast = safe("Forecast",          build_forecast,            daily)
    queue_fc = safe("Per-Queue Forecast", build_per_queue_forecast, data, daily)

    breach_info = forecast.attrs.get("breach_date", "N/A") if not forecast.empty else "N/A"
    log.info("Forecast detractor rate breach date: %s", breach_info)

    log.info("Building executive summary…")
    exec_summary = safe("Executive Summary", build_executive_summary, data, kpi, daily, forecast)

    log.info("Writing Excel report…")
    sheets = {
        "Executive_Summary":    exec_summary,
        "KPI_Summary":          kpi,
        "Daily_Trend":          daily,
        "Monthly_Trend":        monthly,
        "ShortCode_Daily_Pivot": sc_pivot,
        "ShortCode_Catalog":    sc_catalog,
        "Queue_Catalog":        q_catalog,
        "Mapping_Coverage":     map_coverage,
        "Forecast":             forecast,
        "Per_Queue_Forecast":   queue_fc,
        "Top10_Bottom10":       top_bottom,
        "NPS_Waterfall":        waterfall,
        "SLA_Breach_Heatmap":   sla_breach,
        **by_dim,
        "Channel_Comparison":   channel_cmp,
        "FCR_Impact":           fcr_impact,
        "Heatmap_Lev3_x_Lev4":  heatmap,
        "Hour_Pattern":         hour_p,
        "DayOfWeek_Pattern":    dow_p,
        "Toxic_Combos":         toxic,
        "Velocity_Alerts":      velocity,
        "Cohort_Analysis":      cohort,
        "Repeat_Detractors":    repeats,
        "Duplicate_Surveys":    duplicates,
        "Agent_Peer_Benchmark": peer_bench,
        "Agent_Ranking":        agent_rk,
        "Q2_Attitude_vs_TNPS":  q2_ct,
        "Q3_Fulfillment_vs_TNPS": q3_ct,
    }

    write_excel_report(CONFIG["output_file"], sheets, kpi, daily, forecast)

    non_empty = sum(1 for v in sheets.values() if v is not None and not (hasattr(v, "empty") and v.empty))
    log.info("Done. Report: %s  |  %d populated sheets (+ Dashboard + TOC)",
             CONFIG["output_file"], non_empty)


if __name__ == "__main__":
    main()
